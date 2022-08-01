# -*- coding: utf-8 -*-
import os
import time
import webbrowser
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QInputDialog
import pandas as pd
import xlsxwriter as xw
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from win32com.client.gencache import EnsureDispatch
import win32com.client

from Investigation import Enterprise
import re
import requests
import json

class Ui_MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(Ui_MainWindow,self).__init__()
        self.setupUi(self)
        self.retranslateUi(self)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(816, 531)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.scrollArea = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 779, 926))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label_36 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_36.setObjectName("label_36")
        self.horizontalLayout.addWidget(self.label_36)
        self.filetextBrowser = QtWidgets.QTextBrowser(self.scrollAreaWidgetContents)
        self.filetextBrowser.setObjectName("filetextBrowser")
        self.horizontalLayout.addWidget(self.filetextBrowser)
        self.filedirpushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.filedirpushButton.setObjectName("filedirpushButton")
        self.horizontalLayout.addWidget(self.filedirpushButton)
        self.gridLayout_3.addLayout(self.horizontalLayout, 6, 0, 1, 2)
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.label_20 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_20.setObjectName("label_20")
        self.gridLayout.addWidget(self.label_20, 0, 0, 1, 1)
        self.namelineEdit = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        self.namelineEdit.setObjectName("namelineEdit")
        self.gridLayout.addWidget(self.namelineEdit, 0, 1, 1, 1)
        self.label_25 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_25.setObjectName("label_25")
        self.gridLayout.addWidget(self.label_25, 0, 2, 1, 1)
        self.pwdlineEdit = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        self.pwdlineEdit.setObjectName("pwdlineEdit")
        self.gridLayout.addWidget(self.pwdlineEdit, 0, 3, 1, 1)
        self.usercheckBox = QtWidgets.QCheckBox(self.scrollAreaWidgetContents)
        self.usercheckBox.setObjectName("usercheckBox")
        self.gridLayout.addWidget(self.usercheckBox, 0, 4, 1, 1)
        self.label_26 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_26.setObjectName("label_26")
        self.gridLayout.addWidget(self.label_26, 1, 0, 1, 1)
        self.proxieslineEdit = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        self.proxieslineEdit.setObjectName("proxieslineEdit")
        self.gridLayout.addWidget(self.proxieslineEdit, 1, 1, 1, 3)
        self.label_27 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_27.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_27.setObjectName("label_27")
        self.gridLayout.addWidget(self.label_27, 2, 0, 1, 4)
        self.gridLayout_3.addLayout(self.gridLayout, 1, 0, 1, 2)
        self.label_48 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_48.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_48.setObjectName("label_48")
        self.gridLayout_3.addWidget(self.label_48, 10, 1, 1, 1)
        self.gridLayout_9 = QtWidgets.QGridLayout()
        self.gridLayout_9.setObjectName("gridLayout_9")
        self.clear_listslineEdit = QtWidgets.QTextEdit(self.scrollAreaWidgetContents)
        self.clear_listslineEdit.setObjectName("clear_listslineEdit")
        self.gridLayout_9.addWidget(self.clear_listslineEdit, 2, 0, 1, 1)
        self.clear_listscheckBox = QtWidgets.QCheckBox(self.scrollAreaWidgetContents)
        self.clear_listscheckBox.setObjectName("clear_listscheckBox")
        self.gridLayout_9.addWidget(self.clear_listscheckBox, 2, 1, 1, 1)
        self.label_34 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_34.setObjectName("label_34")
        self.gridLayout_9.addWidget(self.label_34, 0, 0, 1, 1)
        self.gridLayout_10 = QtWidgets.QGridLayout()
        self.gridLayout_10.setObjectName("gridLayout_10")
        self.radioButton_save = QtWidgets.QRadioButton(self.scrollAreaWidgetContents)
        self.radioButton_save.setObjectName("radioButton_save")
        self.gridLayout_10.addWidget(self.radioButton_save, 0, 1, 1, 1)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.label_40 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_40.setObjectName("label_40")
        self.horizontalLayout_5.addWidget(self.label_40)
        self.comboBox = QtWidgets.QComboBox(self.scrollAreaWidgetContents)
        self.comboBox.setEnabled(True)
        self.comboBox.setObjectName("comboBox")
        self.horizontalLayout_5.addWidget(self.comboBox)
        self.gridLayout_10.addLayout(self.horizontalLayout_5, 0, 2, 1, 1)
        self.radioButton_clear = QtWidgets.QRadioButton(self.scrollAreaWidgetContents)
        self.radioButton_clear.setChecked(True)
        self.radioButton_clear.setObjectName("radioButton_clear")
        self.gridLayout_10.addWidget(self.radioButton_clear, 0, 0, 1, 1)
        self.label_41 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_41.setObjectName("label_41")
        self.gridLayout_10.addWidget(self.label_41, 0, 3, 1, 1)
        self.label_35 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_35.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_35.setObjectName("label_35")
        self.gridLayout_10.addWidget(self.label_35, 1, 0, 1, 4)
        self.gridLayout_9.addLayout(self.gridLayout_10, 1, 0, 1, 1)
        self.gridLayout_3.addLayout(self.gridLayout_9, 3, 0, 1, 2)
        self.gridLayout_11 = QtWidgets.QGridLayout()
        self.gridLayout_11.setObjectName("gridLayout_11")
        self.savefilepushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.savefilepushButton.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(10)
        sizePolicy.setHeightForWidth(self.savefilepushButton.sizePolicy().hasHeightForWidth())
        self.savefilepushButton.setSizePolicy(sizePolicy)
        self.savefilepushButton.setObjectName("savefilepushButton")
        self.gridLayout_11.addWidget(self.savefilepushButton, 0, 2, 1, 1)
        self.label_43 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_43.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_43.setObjectName("label_43")
        self.gridLayout_11.addWidget(self.label_43, 1, 0, 1, 3)
        self.label_37 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_37.setObjectName("label_37")
        self.gridLayout_11.addWidget(self.label_37, 0, 0, 1, 1)
        self.readfileBrowser = QtWidgets.QTextBrowser(self.scrollAreaWidgetContents)
        self.readfileBrowser.setObjectName("readfileBrowser")
        self.gridLayout_11.addWidget(self.readfileBrowser, 0, 1, 1, 1)
        self.gridLayout_3.addLayout(self.gridLayout_11, 7, 0, 1, 2)
        self.gridLayout_12 = QtWidgets.QGridLayout()
        self.gridLayout_12.setObjectName("gridLayout_12")
        self.readfile2Browser = QtWidgets.QTextBrowser(self.scrollAreaWidgetContents)
        self.readfile2Browser.setObjectName("readfile2Browser")
        self.gridLayout_12.addWidget(self.readfile2Browser, 0, 3, 1, 1)
        self.label_46 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_46.setObjectName("label_46")
        self.gridLayout_12.addWidget(self.label_46, 0, 2, 1, 1)
        self.label_44 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_44.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_44.setObjectName("label_44")
        self.gridLayout_12.addWidget(self.label_44, 1, 2, 1, 2)
        self.savefile2pushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.savefile2pushButton.setObjectName("savefile2pushButton")
        self.gridLayout_12.addWidget(self.savefile2pushButton, 0, 6, 1, 1)
        self.label_45 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_45.setObjectName("label_45")
        self.gridLayout_12.addWidget(self.label_45, 0, 0, 1, 1)
        self.savefile2textEdit = QtWidgets.QTextEdit(self.scrollAreaWidgetContents)
        self.savefile2textEdit.setObjectName("savefile2textEdit")
        self.gridLayout_12.addWidget(self.savefile2textEdit, 0, 1, 1, 1)
        self.gridLayout_3.addLayout(self.gridLayout_12, 8, 0, 1, 2)
        self.gridLayout_7 = QtWidgets.QGridLayout()
        self.gridLayout_7.setObjectName("gridLayout_7")
        self.label_28 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_28.setObjectName("label_28")
        self.gridLayout_7.addWidget(self.label_28, 0, 0, 1, 1)
        self.label_29 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_29.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_29.setObjectName("label_29")
        self.gridLayout_7.addWidget(self.label_29, 1, 0, 1, 2)
        self.industrylineEdit = QtWidgets.QTextEdit(self.scrollAreaWidgetContents)
        self.industrylineEdit.setObjectName("industrylineEdit")
        self.gridLayout_7.addWidget(self.industrylineEdit, 2, 0, 1, 1)
        self.industrycheckBox = QtWidgets.QCheckBox(self.scrollAreaWidgetContents)
        self.industrycheckBox.setObjectName("industrycheckBox")
        self.gridLayout_7.addWidget(self.industrycheckBox, 2, 1, 1, 1)
        self.label_30 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_30.setObjectName("label_30")
        self.gridLayout_7.addWidget(self.label_30, 3, 0, 1, 1)
        self.label_31 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_31.setStyleSheet("color: rgb(255, 0, 0);")
        self.label_31.setObjectName("label_31")
        self.gridLayout_7.addWidget(self.label_31, 4, 0, 1, 2)
        self.keylineEdit = QtWidgets.QTextEdit(self.scrollAreaWidgetContents)
        self.keylineEdit.setObjectName("keylineEdit")
        self.gridLayout_7.addWidget(self.keylineEdit, 5, 0, 1, 1)
        self.keycheckBox = QtWidgets.QCheckBox(self.scrollAreaWidgetContents)
        self.keycheckBox.setObjectName("keycheckBox")
        self.gridLayout_7.addWidget(self.keycheckBox, 5, 1, 1, 1)
        self.gridLayout_3.addLayout(self.gridLayout_7, 2, 0, 1, 2)
        self.gridLayout_13 = QtWidgets.QGridLayout()
        self.gridLayout_13.setObjectName("gridLayout_13")
        self.GopushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.GopushButton.setObjectName("GopushButton")
        self.gridLayout_13.addWidget(self.GopushButton, 0, 0, 1, 1)
        self.openfilepushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.openfilepushButton.setObjectName("openfilepushButton")
        self.gridLayout_13.addWidget(self.openfilepushButton, 0, 1, 1, 1)
        self.pushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout_13.addWidget(self.pushButton, 0, 2, 1, 1)
        self.gridLayout_3.addLayout(self.gridLayout_13, 9, 1, 1, 1)
        self.label_49 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_49.setObjectName("label_49")
        self.gridLayout_3.addWidget(self.label_49, 11, 0, 1, 1)
        self.LogtextBrowser = QtWidgets.QTextBrowser(self.scrollAreaWidgetContents)
        self.LogtextBrowser.setObjectName("LogtextBrowser")
        self.gridLayout_3.addWidget(self.LogtextBrowser, 11, 1, 1, 1)
        self.label = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label.setStyleSheet("font: 40pt \"Adobe Arabic\";")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.gridLayout_3.addWidget(self.label, 0, 0, 1, 2)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout_2.addWidget(self.scrollArea, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 816, 23))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # 用户名
        self.namelineEdit.editingFinished.connect(self.UserName)
        # self.userEdit.setText("wzyt2020")
        # 密码
        self.pwdlineEdit.editingFinished.connect(self.PwdName)
        # self.pwdEdit.setText("WZ20050428")
        #代理ip
        self.proxieslineEdit.editingFinished.connect(self.Proxy)
        # 打开excel
        self.filedirpushButton.clicked.connect(self.Filedir)
        # 企查查招标选择文件夹名
        self.savefilepushButton.clicked.connect(self.saveFBtn)
        # 企查查信息更改文件名
        self.savefile2textEdit.setText("企查查信息2021-11-14")
        self.savefile2pushButton.clicked.connect(self.save2FBtn)
        # 去重的相似度
        rd = ['0.05','0.1','0.15','0.2','0.25','0.30','0.35','0.4','0.45','0.5','0.55','0.6','0.65','0.70','0.75','0.80','0.85','0.9','0.95','1.0']
        self.comboBox.addItems(rd)
        self.comboBox.currentIndexChanged.connect(self.CBox)
        #开始爬取
        self.GopushButton.clicked.connect(self.GoBtn)
        # 打开Excel文件
        self.openfilepushButton.clicked.connect(self.file_open)
        # 点击清除关键词
        self.pushButton.clicked.connect(self.PushButton)
        # 选择清理与保留关键词
        self.radioButton_clear.toggled.connect(self.Radio_clear)

        self.clear_listslineEdit.setPlaceholderText('请输入需清除的关键词')
        self.namelineEdit.setPlaceholderText('请输入企查查账号')
        self.pwdlineEdit.setPlaceholderText('请输入密码')
        self.industrylineEdit.setPlaceholderText('请输入筛选的行业关键词')
        self.keylineEdit.setPlaceholderText('请输入筛选的关键词')
        self.proxieslineEdit.setPlaceholderText('请输入代理ip生成api')
    # 用户名
    def UserName(self):
        self.LogtextBrowser.append("企查查账号：" + self.namelineEdit.text())
        return self.namelineEdit.text()

    # 密码
    def PwdName(self):
        self.LogtextBrowser.append("企查查密码：" + self.pwdlineEdit.text())
        return self.pwdlineEdit.text()
    # 代理ip
    def Proxy(self):
        self.LogtextBrowser.append("代理ip：" + self.proxieslineEdit.text())
        # print(self.proxieslineEdit.text())
        return self.proxieslineEdit.text()
    # 打开excel文件
    def Filedir(self):
        dir_choose1 = QFileDialog.getOpenFileName(self, '选择文件','')
        if dir_choose1 != ('', ''):
            dirfile = list(dir_choose1)[0]
            self.filetextBrowser.append(dirfile)
        else:
            QMessageBox.information(self, "温馨提示", "已取消")
    # 企查查招标更改文件夹名
    def saveFBtn(self):
        dir_choose = QFileDialog.getExistingDirectory(self, '选择文件夹', '')
        if dir_choose != '':
            self.readfileBrowser.setPlainText(dir_choose)
            self.LogtextBrowser.append("文件夹位置：" + self.readfileBrowser.toPlainText())
        else:
            self.readfileBrowser.setPlainText('')
            self.LogtextBrowser.append("文件夹位置：已取消")

    # 企查查信息更改文件名
    def save2FBtn(self):
        dir_choose1 = QFileDialog.getExistingDirectory(self, '另存为', '')
        if dir_choose1 != '':
            self.readfile2Browser.setPlainText(dir_choose1 + "/" +self.savefile2textEdit.toPlainText()+".xlsx")
            self.LogtextBrowser.append("文件位置：" + self.readfile2Browser.toPlainText())
        else:
            self.readfile2Browser.setPlainText('')
            self.LogtextBrowser.append("文件位置：已取消")
    # 选择清理与保留关键词
    def Radio_clear(self):
        if self.radioButton_clear.isChecked():
            self.clear_listslineEdit.setPlaceholderText('请输入需清除的关键词')
            self.LogtextBrowser.append("选择清除关键词：" + self.clear_listslineEdit.toPlainText())
            return 1
        else:
            self.clear_listslineEdit.setPlaceholderText('请输入需保留标题的关键词')
            self.LogtextBrowser.append("选择保留标题的关键词：" + self.clear_listslineEdit.toPlainText())
            return 2
    # 去重的相似度
    def CBox(self):
        return self.comboBox.currentText()
    # 开始爬取
    def GoBtn(self):
        if self.Proxy() != '':
            resp = requests.get(self.Proxy())
            ip = resp.text
            if re.match(r'(?:(?:25[0-5]|2[0-4]\d|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)', ip) is None:
                rr = json.loads(ip)
                hh = rr['data'][0]['ip'] + ":" + str(rr['data'][0]['port'])
            else:
                hh = ip
        else:
            hh = ''
        if QMessageBox.Yes == QMessageBox.information(self, "温馨提示", "准备好吗？", QMessageBox.Yes | QMessageBox.No):
            QMessageBox.information(self, "温馨提醒", "如需要停止运行，请按ctrl+c快捷键", QMessageBox.Yes)
            Ent = Enterprise(hh)
            Ent.Login(self.UserName(), self.PwdName())
            Ent.open_excel(self.filetextBrowser.toPlainText().replace('/','\\'))
            Nfile = self.readfileBrowser.toPlainText().replace('/','\\')+'\\'
            keys_dicts = self.keylineEdit.toPlainText()
            Industry_dicts = self.industrylineEdit.toPlainText()
            clear_lists = self.clear_listslineEdit.toPlainText()
            word_data = Ent.key_word(Nfile, keys_dicts, Industry_dicts, clear_lists, self.CBox(),self.Radio_clear())
            Ent.xw_toExcel(word_data, self.readfile2Browser.toPlainText().replace('/','\\'))
            Ent.Ddel()
            QMessageBox.information(self, "温馨提醒", "完成爬取数据", QMessageBox.Yes)

    # 打开企查查信息文件
    def file_open(self):
        webbrowser.open(self.readfile2Browser.toPlainText())
    def PushButton(self):
        clear_listss, ok = QInputDialog.getText(self, "清除关键词", "请输入清除excel中标题的关键词")
        if ok:
            dir_choose2 = QFileDialog.getOpenFileName(self, '选择文件','')
            if dir_choose2 != ('', ''):
                dirfile = list(dir_choose2)[0]
                fileName2 = dirfile.replace('/', '\\')
                df = pd.read_excel(fileName2, sheet_name='明细')
                data2 = df.values
                ll = []
                for i in range(len(data2)):
                    tiem = {}
                    tiem['项目名称'] = data2[i][0]
                    tiem['名称链接'] = data2[i][1]
                    tiem['发布日期'] = data2[i][2]
                    tiem['省份地区'] = data2[i][3]
                    tiem['信息类型'] = data2[i][4]
                    tiem['招标/采购单位'] = data2[i][5]
                    tiem['中标金额'] = data2[i][6]
                    tiem['关键词'] = data2[i][7]
                    tiem['行业'] = data2[i][8]
                    ll.append(tiem)
                clear_lists = clear_listss.translate(str.maketrans({"，": ",", " ": ""})).split(",")
                listData = []
                for lw in ll:
                    for clear in clear_lists:
                        if clear != '':
                            if lw['项目名称'].find(clear) != -1:
                                listData.append(lw)
                                break
                for lD in listData:
                    for d in ll:
                        try:
                            if lD['项目名称'] == d['项目名称']:
                                ll.remove(lD)
                        except ValueError:
                            pass
                lists_wordsq = ll
                try:
                    workbook1 = xw.Workbook(fileName2)
                    worksheet1 = workbook1.add_worksheet("明细")
                    worksheet1.activate()
                    title1 = ["项目名称", "名称链接", "发布日期", "省份地区", "信息类型", "招标/采购单位", "中标金额", "关键词", "行业"]  # 设置表头
                    worksheet1.write_row('A1', title1)
                    i1 = 2
                    for j1 in range(len(lists_wordsq)):
                        insertData1 = [lists_wordsq[j1]["项目名称"], lists_wordsq[j1]["名称链接"], lists_wordsq[j1]["发布日期"],
                                       lists_wordsq[j1]["省份地区"], lists_wordsq[j1]["信息类型"], lists_wordsq[j1]["招标/采购单位"],
                                       lists_wordsq[j1]["中标金额"], lists_wordsq[j1]["关键词"], lists_wordsq[j1]["行业"]]
                        worksheet1.write_row('A' + str(i1), insertData1)
                        i1 += 1
                    workbook1.close()
                    wbi1 = load_workbook(fileName2)
                    ws1 = wbi1.active
                    alignment1 = Alignment(horizontal='left', vertical='top', text_rotation=0)
                    rows1 = ws1.max_row
                    for rws1 in range(1, rows1 + 1):
                        ws1["A" + str(rws1)].alignment = alignment1
                        ws1["B" + str(rws1)].alignment = alignment1
                        ws1["C" + str(rws1)].alignment = alignment1
                        ws1["D" + str(rws1)].alignment = alignment1
                        ws1["E" + str(rws1)].alignment = alignment1
                        ws1["F" + str(rws1)].alignment = alignment1
                        ws1["G" + str(rws1)].alignment = alignment1
                        ws1["H" + str(rws1)].alignment = alignment1
                        ws1["I" + str(rws1)].alignment = alignment1
                    ws1.column_dimensions['A'].width = 70
                    ws1.column_dimensions['B'].width = 50
                    ws1.column_dimensions['C'].width = 10
                    ws1.column_dimensions['D'].width = 8
                    ws1.column_dimensions['E'].width = 8
                    ws1.column_dimensions['F'].width = 30
                    ws1.column_dimensions['G'].width = 20
                    wbi1.save(fileName2)
                    wbi1.close()
                    print("完成写入Excel")
                    time.sleep(0.5)
                    print("正在操作数据透视表")
                except:
                    self.LogtextBrowser.append('请关闭当前文件已经开了')
                try:
                    xlApp = EnsureDispatch('Excel.Application')
                    win32c = win32com.client.constants
                    wbwin1 = xlApp.Workbooks.Open(fileName2)
                    St1 = wbwin1.Worksheets("明细")
                    Psr1 = St1.Range(St1.Cells(1, 1), St1.Cells(St1.UsedRange.Rows.Count, St1.UsedRange.Columns.Count))
                    Psr1.Select()
                    St11 = wbwin1.Worksheets.Add()
                    St11.Name = '分类'
                    PivotC1 = wbwin1.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=Psr1,
                                                          Version=win32c.xlPivotTableVersion14)
                    pT1 = PivotC1.CreatePivotTable(TableDestination=St11.Range(St11.Cells(1, 1), St11.Cells(1, 1)),
                                                   TableName='数据透视表', DefaultVersion=win32c.xlPivotTableVersion14)
                    pT1.RowAxisLayout(win32c.xlOutlineRow)
                    pT1.PivotFields('行业').Orientation = win32c.xlRowField
                    pT1.PivotFields('关键词').Orientation = win32c.xlRowField
                    pT1.PivotFields('省份地区').Orientation = win32c.xlRowField
                    pT1.PivotFields('项目名称').Orientation = win32c.xlRowField
                    pT1.PivotFields('行业').Position = 1
                    pT1.PivotFields('关键词').Position = 2
                    pT1.PivotFields('省份地区').Position = 3
                    pT1.PivotFields('项目名称').Position = 4
                    t = pT1.PivotFields('行业').DataRange.Columns
                    n = 0
                    for dd in t.Value:
                        if dd[0] is not None:
                            n += 1
                    pT1.PivotFields('行业').PivotItems('其他').Position = n
                    wbwin1.SaveAs(fileName2)
                    xlApp.Quit()
                    self.LogtextBrowser.append('完成清除关键词')
                except:
                    self.LogtextBrowser.append('Excel操作数据透视表有误，请手动操作一下')
            else:
                QMessageBox.information(self, "温馨提示", "已取消")
        else:
            QMessageBox.information(self, "温馨提示", "已取消")

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label_36.setText(_translate("MainWindow", "选择企查查供应商名称Excel文件:"))
        self.filedirpushButton.setText(_translate("MainWindow", "选择文件"))
        self.label_20.setText(_translate("MainWindow", "企查查账号："))
        self.label_25.setText(_translate("MainWindow", "企查查密码："))
        self.usercheckBox.setText(_translate("MainWindow", "记住用户"))
        self.label_26.setText(_translate("MainWindow", "代理ip："))
        self.label_27.setText(_translate("MainWindow", "注：如使用代理ip生成api链接，就输入api链接，如果不使用，请清空一下，否则报错"))
        self.label_48.setText(_translate("MainWindow", "（注：打开企查查文件前，要等到爬取结束再打开文件即可"))
        self.clear_listscheckBox.setText(_translate("MainWindow", "记住"))
        self.label_34.setText(_translate("MainWindow", "选择保留的标题关键词或者需清理的关键词：（需保存，请打勾记住）"))
        self.radioButton_save.setText(_translate("MainWindow", "保留的标题关键词"))
        self.label_40.setText(_translate("MainWindow", "相似度："))
        self.radioButton_clear.setText(_translate("MainWindow", "需清理的关键词"))
        self.label_41.setText(_translate("MainWindow", "需招标信息去重，请选择一下相似度"))
        self.label_35.setText(_translate("MainWindow", "（注：输入关键词，多个关键词组合使用（词和词之间使用逗号隔开），例如：单一,租赁，最后不要加“，”）"))
        self.savefilepushButton.setText(_translate("MainWindow", "选择文件夹"))
        self.label_43.setText(
            _translate("MainWindow", "（注：这是把招标文件保存文件夹里面。保存招标信息的excel之前，需要选择文件夹，或者最好在同一企查查信息文件下创建文件夹）"))
        self.label_37.setText(_translate("MainWindow", "招中标文件夹路径："))
        self.label_46.setText(_translate("MainWindow", "结果存放路径："))
        self.label_44.setText(_translate("MainWindow", "（注：这是保存企查查信息的excel）"))
        self.savefile2pushButton.setText(_translate("MainWindow", "另存文件"))
        self.label_45.setText(_translate("MainWindow", "自定义文件名："))
        self.label_28.setText(_translate("MainWindow", "筛选的行业关键词：（需保存，请打勾记住）"))
        self.label_29.setText(_translate("MainWindow", "（注：输入关键词，多个关键词组合使用（词和词之间使用逗号隔开），例如：公积金,档案馆，最后不要加“，”）"))
        self.industrycheckBox.setText(_translate("MainWindow", "记住"))
        self.label_30.setText(_translate("MainWindow", "筛选的关键词：（需保存，请打勾记住）"))
        self.label_31.setText(_translate("MainWindow", "（注：输入关键词，多个关键词组合使用（词和词之间使用逗号隔开），例如：运维服务,数字化，最后不要加“，”）"))
        self.keycheckBox.setText(_translate("MainWindow", "记住"))
        self.GopushButton.setText(_translate("MainWindow", "开始爬取"))
        self.openfilepushButton.setText(_translate("MainWindow", "打开企查查文件"))
        self.pushButton.setText(_translate("MainWindow", "已有Excel文件保留的关键词，需清除关键词"))
        self.label_49.setText(_translate("MainWindow", "日志："))
        self.label.setText(_translate("MainWindow", "企查查公司信息抓取"))

if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())