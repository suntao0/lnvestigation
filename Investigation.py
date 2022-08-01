# -*- coding: utf-8 -*-
import difflib
import json
import time
import re

import requests
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import  Alignment
import selenium.webdriver as wb
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
import xlsxwriter as xw
from win32com.client.gencache import EnsureDispatch
import win32com.client
import os
class Enterprise:
    def __init__(self,PROXY):
        self.options = Options()
        # 禁止加载图片
        # prefs = {"profile.managed_default_content_settings.images": 2}
        prefs={"":""}
        # 避免密码提示框的弹出
        prefs["credentials_enable_service"] = False
        prefs["profile.password_manager_enabled"] = False
        self.options.add_experimental_option("prefs", prefs)
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        # 不能隐藏浏览器，会被检测到，要求登入账号
        # options.add_argument('--headless')
        self.options.add_argument('--incognito')  # 隐身模式（无痕模式）
        self.options.add_argument('--disable-gpu')  # 禁用gpu，解决一些莫名的问题
        # self.options.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片, 提升速度
        self.options.add_argument('--disable-infobars')  # 禁用浏览器正在被自动化程序控制的提示
        self.options.add_argument('--start-maximized')
        self.options.add_argument('--user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.93 Safari/537.36"')
        # 设置代理
        self.options.add_argument('--proxy-server=' + PROXY)

        # 隐藏 正在受到自动软件的控制 这几个字
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option('useAutomationExtension', False)
        self.options.add_argument('log-level=3')
        # 设置无头模式
        self.options.headless = False
        self.driver = wb.Chrome(options=self.options)
        # self.driver.maximize_window() # 最大窗口
        # 为了避免页面空白，故加俩个get
        print("进入登录页面")
        self.driver.get('https://www.qcc.com/weblogin')
        self.driver.get('https://www.qcc.com/weblogin')
        time.sleep(1)
    # 捕获异常
    def NodeExists(self,xpath):
        try:
            self.driver.find_element_by_xpath(xpath)
            return True
        except:
            return False
    def Login(self, username, password):
        time.sleep(1)
        self.driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[2]/div/div[2]').click()
        try:
            if self.NodeExists('//*[@id="nc_1_n1z"]/span'):
                if re.match(r"^1[35678]\d{9}$", username):
                    try:
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[1]/div[2]').click()
                        # # 找到账号输入框
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[3]/form/div[1]/input').send_keys(username)
                        # # 找到密码输入框
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[3]/form/div[2]/input').send_keys(password)
                        start = self.driver.find_element_by_xpath('//*[@id="nc_1_n1z"]/span')
                        # 长按拖拽
                        action = ActionChains(self.driver)
                        action.click_and_hold(start)
                        action.drag_and_drop_by_offset(start, 308, 0).perform()
                        time.sleep(2)
                        # 点击登录
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[3]/form/div[4]/button').click()
                        time.sleep(3)
                    except:
                        self.driver.quit()
                        print("您输入的账号和密码有误")
                else:
                    self.driver.quit()
                    print('您输入的手机号码格式错误')
            else:
                if re.match(r"^1[35678]\d{9}$", username):
                    try:
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[1]/div[2]').click()
                        # # 找到账号输入框
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[3]/form/div[1]/input').send_keys(username)
                        # # 找到密码输入框
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[3]/form/div[2]/input').send_keys(password)
                        self.driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/div[2]/div[3]/form/div[4]/button').click()
                        time.sleep(3)
                    except:
                        self.driver.quit()
                        print("您输入的账号和密码有误")
                else:
                    self.driver.quit()
                    print('您输入的手机号码格式错误')
            print("登录成功，准备进入企查查页面")
        except:
            self.driver.quit()
            print("登录页面异常，请检查是否弹出对话框，请手动操作一下")
    def open_excel(self,filename):
        wb = load_workbook(filename)
        ws = wb.active
        global lis
        lis = []
        row = ws.max_row
        for row in range(2, row + 1):
            if ws.cell(row=row, column=1).value != None:
                lis.append(ws.cell(row=row, column=1).value)
    def key_word(self,Newfile,k_dicts,I_dicts,c_lists,rds,save_clear):
        keys_dicts     = k_dicts.translate(str.maketrans({"，":","," ":""})).split(",")
        Industry_dicts = I_dicts.translate(str.maketrans({"，":","," ":""})).split(",")
        clear_lists    = c_lists.translate(str.maketrans({"，":","," ":""})).split(",")
        lists_data = []
        for i in lis:
            print(f"正在选择{i}")
            self.driver.get(f'https://www.qcc.com/web/search?key={i}')
            # 判断公司全称存在不存在
            print(f"正在爬取{i}信息")
            if self.NodeExists("//tr[@class='frtrt tsd0']//td[3]"):
                dd = self.driver.find_element_by_xpath("//tr[@class='frtrt tsd0']//td[3]").text
                sre = dd.replace('\n', ' ').split(" ")
                item = {}
                item['供应商名称'] = self.driver.find_element_by_xpath("//tr[@class='frtrt tsd0']//a[@class='title copy-value']//span").text
                try:
                    item['企业法人'] = self.driver.find_element_by_xpath("//table[@class='ntable ntable-list']/tr[1]/td[3]//span[@class='val']").text
                except ValueError:
                    item['企业法人'] = '无'
                try:
                    item['注册资金'] = sre[sre.index('注册资本：') + 1]
                except ValueError:
                    item['注册资金'] = '无'
                try:
                    item['统一社会信用代码'] = sre[sre.index('统一社会信用代码：') + 1]
                except ValueError:
                    item['统一社会信用代码'] = '无'
                try:
                    item['成立日期'] = sre[sre.index('成立日期：') + 1]
                except ValueError:
                    item['成立日期'] = '无'
                try:
                    item['电话'] = sre[sre.index('电话：') + 1]
                except ValueError:
                    item['电话'] = '无'
                try:
                    item['企业邮箱'] = sre[sre.index('邮箱：') + 1]
                except ValueError:
                    item['企业邮箱'] = '无'
                try:
                    item['企业官网'] = sre[sre.index('官网：') + 1]
                except ValueError:
                    item['企业官网'] = '无'
                self.driver.find_element_by_xpath("//tr[@class='frtrt tsd0']//a[@class='title copy-value']//span").click()
                self.driver.switch_to.window(self.driver.window_handles[-1])
                try:
                    item['企业简介'] = self.driver.find_element_by_xpath("//div[@class='rline extend-text']/span[@class='f']/span[@class='val']").text
                except:
                    item['企业简介'] = '暂无'
                if self.NodeExists("//body/div/div[@class='company-detail']/div[@class='company-header']/div[@class='container']/div[@class='nheader']/div[@class='infos clearfix']/div[@class='content']/div[@class='contact-info']/div[3]/span[1]/a[1]"):
                    inc_full = self.driver.find_element_by_xpath("//body/div/div[@class='company-detail']/div[@class='company-header']/div[@class='container']/div[@class='nheader']/div[@class='infos clearfix']/div[@class='content']/div[@class='contact-info']/div[3]/span[1]/a[1]").text
                    if inc_full[0:2] == "更多":
                        self.driver.find_element_by_xpath("//body/div/div[@class='company-detail']/div[@class='company-header']/div[@class='container']/div[@class='nheader']/div[@class='infos clearfix']/div[@class='content']/div[@class='contact-info']/div[3]/span[1]/a[1]").click()
                        time.sleep(1)
                        item['更多邮箱'] = self.driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='app-nmodal modal fade in']/div[@class='modal-dialog nmodal']/div[@class='modal-content']/div[@class='modal-body']/div/div[@class='more-list']/table").text
                        time.sleep(1)
                        element=self.driver.find_element_by_xpath("/html/body[@class='modal-open']/div[@class='app-nmodal modal fade in']/div[@class='modal-dialog nmodal']/div[@class='modal-content']/div[@class='modal-header']/a[@class='nclose']")
                        self.driver.execute_script("arguments[0].click();", element)
                    else:
                        item['更多邮箱'] = '无'
                else:
                    item['更多邮箱'] = '无'
                time.sleep(1)
                if self.NodeExists("/html/body/div[1]/div[2]/div[1]/div/div[1]/div[2]/div[2]/div[3]/div[2]/span[1]/a[1]"):
                    inc_phone = self.driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[1]/div/div[1]/div[2]/div[2]/div[3]/div[2]/span[1]/a[1]").text
                    if inc_phone[:2] == "更多":
                        self.driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[1]/div/div[1]/div[2]/div[2]/div[3]/div[2]/span[1]/a[1]").click()
                        time.sleep(1)
                        if self.NodeExists("//div[@class='app-nmodal modal fade in']//div[@class='modal-body']"):
                            ite = self.driver.find_element_by_xpath("//div[@class='app-nmodal modal fade in']//div[@class='modal-body']").text
                            time.sleep(1)
                            item['更多电话'] = re.sub('同电话企业 \d+', '', ite)
                            # 关闭弹出对话框
                            eleme=self.driver.find_element_by_xpath("//div[@class='app-nmodal modal fade in']//a[@class='nclose']")
                            self.driver.execute_script("arguments[0].click();", eleme)
                        else:
                            # 关闭弹出对话框
                            el = self.driver.find_element_by_xpath("//button[@class='close']//span[1]")
                            self.driver.execute_script("arguments[0].click();", el)
                            item['更多电话'] = '您不是会员，无法查看'
                    else:
                        item['更多电话'] = '无'
                else:
                    item['更多电话'] = '无'
                time.sleep(1)
                print(f"完成爬取{i}信息")
                # 招标信息
                time.sleep(1)
                # 判断经营信息存在不存在
                if self.NodeExists("//div[@class='nav-head']"):
                    nav = self.driver.find_element_by_xpath("//div[@class='nav-head']").text
                    for navj in range(len(nav.split('\n'))):
                        if self.NodeExists(f"//div[@class='nav-head']//a[{navj + 1}]"):
                            if self.driver.find_element_by_xpath(f"//div[@class='nav-head']//a[{navj + 1}]").text[:4] == '经营信息':
                                el = self.driver.find_element_by_xpath(f"//div[@class='nav-head']//a[{navj + 1}]")
                                self.driver.execute_script("arguments[0].click();", el)
                                time.sleep(2)
                                # 判断招投标的定位元素，如数量为0,则招标无数据
                                if self.NodeExists("//section[@id='tenderlist']//div//h3"):
                                    item['文件路径'] = f'=HYPERLINK("{Newfile + i}.xlsx","{i}")'
                                    time.sleep(0.5)
                                    pn1l = int(self.driver.find_element_by_xpath("//section[@id='tenderlist']//div//span[1]").text)
                                    time.sleep(1)
                                    pl = self.driver.find_element_by_xpath('//*[@id="tenderlist"]/div[2]/table')
                                    rows = len(pl.find_elements_by_tag_name('tr')) - 1
                                    pr = pn1l/10
                                    wd = str(pr)
                                    # 判断总招标数量/10  >  1,则点击下一页
                                    lists_wordsq = []
                                    if pr > 1:
                                        n3 = 1
                                        if pn1l % 10 != 0:
                                            n4 = int(wd[:-2]) + 1
                                        else:
                                            n4 = int(wd[:-2])
                                        while True:
                                            if self.NodeExists("//section[@id='tenderlist']//a[contains(text(),'>')]"):
                                                print(f"正在爬取第{n3}页，共{n4}页")
                                                # 总招标数据列表
                                                plp = self.driver.find_element_by_xpath('//*[@id="tenderlist"]/div[2]/table')
                                                rowsp = len(plp.find_elements_by_tag_name('tr')) - 1
                                                for j in range(2, rowsp + 2):
                                                    item_wordq = {}
                                                    item_wordq['项目名称'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[2]/span').text
                                                    item_wordq['名称链接'] = self.driver.find_element_by_xpath(f"//*[@id='tenderlist']//tr[{j}]//td[8]//span[1]/a").get_attribute("href")
                                                    item_wordq['发布日期'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[3]').text
                                                    item_wordq['省份地区'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[4]/span').text
                                                    item_wordq['信息类型'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[5]/span').text
                                                    item_wordq['招标/采购单位'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[6]/span').text
                                                    item_wordq['中标金额'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[7]/span').text
                                                    def Keys7(word):
                                                        for str1 in keys_dicts:
                                                            if str1 !='':
                                                                if word.find(str1) != -1:
                                                                    return str1
                                                        return "其他"
                                                    item_wordq["关键词"] = Keys7(item_wordq['项目名称'])
                                                    def Industry7(word):
                                                        for str11 in Industry_dicts:
                                                            if str11 !='':
                                                                if word.find(str11) != -1:
                                                                    return str11
                                                        return "其他"
                                                    item_wordq["行业"] = Industry7(item_wordq['项目名称'])
                                                    lists_wordsq.append(item_wordq)
                                                elm = self.driver.find_element_by_xpath("//section[@id='tenderlist']//a[contains(text(),'>')]")
                                                self.driver.execute_script("arguments[0].click();", elm)
                                                time.sleep(5)
                                            else:
                                                print(f"正在爬取第{n4}页，共{n4}页")
                                                # 总招标数据列表
                                                plp = self.driver.find_element_by_xpath('//*[@id="tenderlist"]/div[2]/table')
                                                rowsp = len(plp.find_elements_by_tag_name('tr')) - 1
                                                for j in range(2, rowsp + 2):
                                                    item_wordq = {}
                                                    item_wordq['项目名称'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[2]/span').text
                                                    item_wordq['名称链接'] = self.driver.find_element_by_xpath(f"//*[@id='tenderlist']//tr[{j}]//td[8]//span[1]/a").get_attribute("href")
                                                    item_wordq['发布日期'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[3]').text
                                                    item_wordq['省份地区'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[4]/span').text
                                                    item_wordq['信息类型'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[5]/span').text
                                                    item_wordq['招标/采购单位'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[6]/span').text
                                                    item_wordq['中标金额'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[7]/span').text
                                                    def Keys8(word):
                                                        for str1 in keys_dicts:
                                                            if str1 !='':
                                                                if word.find(str1) != -1:
                                                                    return str1
                                                        return "其他"
                                                    item_wordq["关键词"] = Keys8(item_wordq['项目名称'])
                                                    def Industry8(word):
                                                        for str11 in Industry_dicts:
                                                            if str11 !='':
                                                                if word.find(str11) != -1:
                                                                    return str11
                                                        return "其他"
                                                    item_wordq["行业"] = Industry8(item_wordq['项目名称'])
                                                    lists_wordsq.append(item_wordq)
                                                break
                                            if n3 == n4:
                                                break
                                            n3 += 1
                                    else:
                                        print("正在爬取第1页，共1页")
                                        for j in range(2, rows + 2):
                                            item_wordq = {}
                                            item_wordq['项目名称'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[2]/span').text
                                            item_wordq['名称链接'] = self.driver.find_element_by_xpath(f"//*[@id='tenderlist']//tr[{j}]//td[8]//span[1]/a").get_attribute("href")
                                            item_wordq['发布日期'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[3]').text
                                            item_wordq['省份地区'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[4]/span').text
                                            item_wordq['信息类型'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[5]/span').text
                                            item_wordq['招标/采购单位'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[6]/span').text
                                            item_wordq['中标金额'] = self.driver.find_element_by_xpath(f'//*[@id="tenderlist"]/div[2]/table/tr[{j}]/td[7]/span').text
                                            def Keys1(word1):
                                                for str1 in keys_dicts:
                                                    if str1 !='':
                                                        if word1.find(str1) != -1:
                                                            return str1
                                                return "其他"
                                            item_wordq["关键词"] = Keys1(item_wordq['项目名称'])
                                            def Industry1(word):
                                                for str11 in Industry_dicts:
                                                    if str11 != '':
                                                        if word.find(str11) != -1:
                                                            return str11
                                                return "其他"
                                            item_wordq["行业"] = Industry1(item_wordq['项目名称'])
                                            lists_wordsq.append(item_wordq)
                                    print("完成爬取招标数据")
                                    listData = []
                                    if clear_lists != ['']:
                                        if save_clear == 1:
                                            print("正在清除关键词")
                                            for lw in lists_wordsq:
                                                for clear in clear_lists:
                                                    if clear != '':
                                                        if lw['项目名称'].find(clear) != -1:
                                                            listData.append(lw)
                                                            break
                                            for lD in listData:
                                                for lwq in lists_wordsq:
                                                    try:
                                                        if lD['项目名称'] == lwq['项目名称']:
                                                            lists_wordsq.remove(lD)
                                                    except ValueError:
                                                        pass
                                            print("完成清除关键词")
                                        else:
                                            print("正在保留标题的关键词")
                                            for lw in lists_wordsq:
                                                for clear in clear_lists:
                                                    if clear != '':
                                                        if lw['项目名称'].find(clear) != -1:
                                                            listData.append(lw)
                                                            break
                                            print("完成保留标题的关键词")
                                            lists_wordsq = listData
                                    print("正在将相似度为高去重")
                                    # 数据去重，相似度设为大于rds
                                    # 相邻相似度
                                    ldw = []
                                    for il in range(len(lists_wordsq) - 1):
                                        rd = round(difflib.SequenceMatcher(None, lists_wordsq[il]['项目名称'],lists_wordsq[il + 1]['项目名称']).quick_ratio(), 2)
                                        if rd >= float(rds):
                                            ldw.append(lists_wordsq[il])
                                    for jjj in ldw:
                                        lists_wordsq.remove(jjj)
                                    print("完成将相似度为高去重")
                                    try:
                                        print("正在写入Excel")
                                        try:
                                            workbook1 = xw.Workbook(Newfile+f"1{i}.xlsx")
                                            worksheet1 = workbook1.add_worksheet("明细")
                                            worksheet1.activate()
                                            title1 = ["项目名称", "名称链接", "发布日期", "省份地区", "信息类型", "招标/采购单位", "中标金额","关键词","行业"]  # 设置表头
                                            worksheet1.write_row('A1', title1)
                                            i1 = 2
                                            for j1 in range(len(lists_wordsq)):
                                                insertData1 = [lists_wordsq[j1]["项目名称"], lists_wordsq[j1]["名称链接"], lists_wordsq[j1]["发布日期"],lists_wordsq[j1]["省份地区"], lists_wordsq[j1]["信息类型"],lists_wordsq[j1]["招标/采购单位"], lists_wordsq[j1]["中标金额"],lists_wordsq[j1]["关键词"],lists_wordsq[j1]["行业"]]
                                                worksheet1.write_row('A' + str(i1), insertData1)
                                                i1 += 1
                                            workbook1.close()
                                            wbi1 = load_workbook(Newfile+f"1{i}.xlsx")
                                            ws1 = wbi1.active
                                            alignment1 = Alignment(horizontal='left', vertical='top', text_rotation=0)
                                            rows1 = ws1.max_row
                                            for rws1 in range(1,rows1+1):
                                                ws1["A" + str(rws1)].alignment = alignment1
                                                ws1["B" + str(rws1)].alignment = alignment1
                                                ws1["C" + str(rws1)].alignment = alignment1
                                                ws1["D" + str(rws1)].alignment = alignment1
                                                ws1["E" + str(rws1)].alignment = alignment1
                                                ws1["F" + str(rws1)].alignment = alignment1
                                                ws1["G" + str(rws1)].alignment = alignment1
                                                ws1["H" + str(rws1)].alignment = alignment1
                                                ws1["I" + str(rws1)].alignment = alignment1
                                            ws1.column_dimensions['A'].width = 70
                                            ws1.column_dimensions['B'].width = 50
                                            ws1.column_dimensions['C'].width = 10
                                            ws1.column_dimensions['D'].width = 8
                                            ws1.column_dimensions['E'].width = 8
                                            ws1.column_dimensions['F'].width = 30
                                            ws1.column_dimensions['G'].width = 20
                                            wbi1.save(Newfile+f"1{i}.xlsx")
                                            wbi1.close()
                                            print("完成写入Excel")
                                            time.sleep(0.5)
                                            print("正在操作数据透视表")
                                        except xw.exceptions.FileCreateError:
                                            print("请创建文件夹")
                                            self.driver.quit()
                                        try:
                                            xlApp = EnsureDispatch('Excel.Application')
                                            win32c = win32com.client.constants
                                            wbwin1 = xlApp.Workbooks.Open(Newfile+f"1{i}.xlsx")
                                            St1 = wbwin1.Worksheets("明细")
                                            Psr1 = St1.Range(St1.Cells(1, 1),St1.Cells(St1.UsedRange.Rows.Count, St1.UsedRange.Columns.Count))
                                            Psr1.Select()
                                            St11 = wbwin1.Worksheets.Add()
                                            St11.Name = '分类'
                                            PivotC1 = wbwin1.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=Psr1,Version=win32c.xlPivotTableVersion14)
                                            pT1 = PivotC1.CreatePivotTable(TableDestination=St11.Range(St11.Cells(1, 1), St11.Cells(1, 1)), TableName='数据透视表',DefaultVersion=win32c.xlPivotTableVersion14)
                                            pT1.RowAxisLayout(win32c.xlOutlineRow)
                                            pT1.PivotFields('行业').Orientation = win32c.xlRowField
                                            pT1.PivotFields('关键词').Orientation = win32c.xlRowField
                                            pT1.PivotFields('省份地区').Orientation = win32c.xlRowField
                                            pT1.PivotFields('项目名称').Orientation = win32c.xlRowField
                                            pT1.PivotFields('行业').Position = 1
                                            pT1.PivotFields('关键词').Position = 2
                                            pT1.PivotFields('省份地区').Position = 3
                                            pT1.PivotFields('项目名称').Position = 4
                                            t = pT1.PivotFields('行业').DataRange.Columns
                                            n = 0
                                            for dd in t.Value:
                                                if dd[0] is not None:
                                                    n += 1
                                            pT1.PivotFields('行业').PivotItems('其他').Position = n
                                            wbwin1.SaveAs(Newfile+f"{i}.xlsx")
                                            xlApp.Quit()
                                            os.remove(Newfile+f"1{i}.xlsx")
                                            print("完成操作数据透视表")
                                        except:
                                            print(f"{i}.xlsx数据透视表版本出问题，请手动操作数据透视表")
                                        print(f'{i}.xlsx已保存')
                                    except PermissionError:
                                        print("当前文件正在打开，请关闭当前文件")
                                else:
                                    item['文件路径'] = '招标无数据'
                                    print(f'因招标无数据，已删除{i}.xlsx')
                else:
                    item['文件路径'] = '招标无数据'
                    print(f'因招标无数据，已删除{i}.xlsx')

                self.driver.close()
                self.driver.switch_to.window(self.driver.window_handles[-1])
                lists_data.append(item)
            else:
                print(f'未找到{i}')
                itemw = {}
                itemw['供应商名称'] = i
                itemw['企业法人'] = '未找到公司名称'
                lists_data.append(itemw)
        return lists_data
    def xw_toExcel(self, data, fileName):
        print("正在将企查查信息写入Excel")
        try:
            workbook = xw.Workbook(fileName)
            worksheet = workbook.add_worksheet("企查查信息")
            title = ["供应商名称", "企业法人", "注册资金", "统一社会信用代码","成立日期", "电话", "企业邮箱", "企业官网", "企业简介", "更多邮箱","更多电话","采招标数据(外部文件)，蓝色表示超链接"]
            worksheet.write_row('A1', title)
            i = 2
            for j in range(len(data)):
                if data[j]['企业法人'] != '未找到公司名称':
                    insertData = [data[j]["供应商名称"], data[j]["企业法人"], data[j]["注册资金"], data[j]["统一社会信用代码"],data[j]["成立日期"], data[j]["电话"],data[j]["企业邮箱"], data[j]["企业官网"], data[j]["企业简介"], data[j]["更多邮箱"], data[j]["更多电话"],data[j]["文件路径"]]
                else:
                    insertData=[data[j]["供应商名称"], data[j]["企业法人"]]
                row = 'A' + str(i)
                worksheet.write_row(row, insertData)
                i += 1
            workbook.close()
            wb9 = load_workbook(fileName)
            ws9 = wb9.active
            cols = ws9.max_column
            for col in range(1, cols + 1):
                ws9.cell(row=1, column=col).font = Font(bold=True)
            font1 = Font(color="ff0000")
            fille2 = PatternFill("solid",fgColor="92d050")
            fille = PatternFill("solid", fgColor="FFBB02")
            rowsqq = ws9.max_row
            for i in range(2, rowsqq + 1):
                one_cell = ws9.cell(row=i, column=2)
                if one_cell.value == "未找到公司名称":
                    ws9.cell(row=i, column=2).fill = fille
                    ws9.cell(row=i, column=2).font = font1
                    ws9.cell(row=i, column=1).font = font1
                    ws9.cell(row=i, column=1).fill = fille2
                    ws9.merge_cells(f'B{i}:L{i}')
                    ws9[f'B{i}'] = '未找到公司名称'
            ws9.column_dimensions['A'].width = 35
            ws9.column_dimensions['B'].width = 15
            ws9.column_dimensions['C'].width = 20
            ws9.column_dimensions['D'].width = 30
            ws9.column_dimensions['E'].width = 15
            ws9.column_dimensions['F'].width = 25
            ws9.column_dimensions['G'].width = 26
            ws9.column_dimensions['H'].width = 26
            ws9.column_dimensions['I'].width = 30
            ws9.column_dimensions['J'].width = 40
            ws9.column_dimensions['K'].width = 40
            ws9.column_dimensions['L'].width = 38
            rows = ws9.max_row
            font_false = Font(color="1E88E5")
            for i in range(2, rows + 1):
                alignment9 = Alignment(horizontal='left', vertical='top', text_rotation=0)
                ws9["A" + str(i)].alignment = alignment9
                ws9["B" + str(i)].alignment = alignment9
                ws9["C" + str(i)].alignment = alignment9
                ws9["D" + str(i)].alignment = alignment9
                ws9["E" + str(i)].alignment = alignment9
                ws9["F" + str(i)].alignment = alignment9
                ws9["G" + str(i)].alignment = alignment9
                ws9["H" + str(i)].alignment = alignment9
                ws9["I" + str(i)].alignment = alignment9
                ws9["J" + str(i)].alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
                ws9["K"+  str(i)].alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
                ws9["L" + str(i)].alignment = alignment9

                file_cell = ws9.cell(row=i, column=12)
                if file_cell.value !='招标无数据':
                    ws9.cell(row=i, column=12).font = font_false
                else:
                    ws9.cell(row=i, column=12).font = font1
            wb9.save(fileName)
            wb9.close()
            print("完成将企查查信息写入Excel")
        except PermissionError:
            print("当前文件正在打开，请关闭当前文件")

    def Ddel(self):
        self.driver.quit()
        print("浏览器已退出")
if __name__ == '__main__':
    # 测试各个模块
    Proxy =''
    if Proxy != '':
        ip = Proxy
        if re.match(r'(?:(?:25[0-5]|2[0-4]\d|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)', ip) is None:
            rr = json.loads(ip)
            hh = rr['data'][0]['ip'] + ":" + str(rr['data'][0]['port'])
        else:
            hh = ip
    else:
        hh = ''
    Ent = Enterprise(hh)
    Ent.Login('18656965923','zxcv1234')
    # Ent.open_excel('C:\\Users\\admin\\Desktop\\工作簿1.xlsx')